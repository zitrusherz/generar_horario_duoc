import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Conexión a la base de datos 'horarios.db'
conn = sqlite3.connect('horarios.db')

# Combinaciones de cursos proporcionadas
combinations = [
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-020D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-020D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-020D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-002D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-017D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-017D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-016D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-016D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-021D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-021D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-016D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-016D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-017D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-017D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-020D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-003D,FCE1100-020D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-021D,INU2101-014D,MDY2131-002D,PLC2101-022D,APY4461-004D",
    "MAT4140-005D,PGY4121-006D,CSY4111-007D,ASY4131-002D,FCE1100-021D,INU2101-015D,MDY2131-002D,PLC2101-022D,APY4461-004D"
]






# Crear un nuevo libro de Excel
wb = Workbook()
wb.remove(wb.active)  # Eliminar la hoja activa por defecto

# Rango de tiempo de 40 minutos desde 08:30 a 18:50
time_slots = pd.date_range("08:30", "18:50", freq="40min").strftime('%H:%M').tolist()
days = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

for idx, comb in enumerate(combinations, start=1):
    cursos = comb.split(',')

    # Consultar los datos necesarios de las tablas Horario, Seccion y Curso
    query = """
    SELECT s.codigo_seccion, s.nombre_profesor, c.nombre AS nombre_curso, h.dia, h.inicio, h.fin
    FROM Horario h
    JOIN Seccion s ON h.id_seccion = s.id
    JOIN Curso c ON s.id_curso = c.id
    WHERE s.codigo_seccion IN ({})
    """.format(",".join(f"'{curso}'" for curso in cursos))

    df = pd.read_sql_query(query, conn)

    # Convertir los tiempos a cadenas de texto en el formato 'HH:MM'
    df['inicio'] = pd.to_datetime(df['inicio'], format='%H:%M:%S', errors='coerce').dt.strftime('%H:%M')
    df['fin'] = pd.to_datetime(df['fin'], format='%H:%M:%S', errors='coerce').dt.strftime('%H:%M')

    # Reemplazar valores NaT (errores de conversión) por un string indicando el error
    df['inicio'] = df['inicio'].fillna('Formato inválido')
    df['fin'] = df['fin'].fillna('Formato inválido')

    # Crear una nueva hoja para la combinación actual
    ws = wb.create_sheet(title=f'Combinacion_{idx}')

    # Escribir los encabezados
    ws.append([""] + days)
    for time in time_slots:
        ws.append([time] + [""] * len(days))

    # Llenar la hoja con los datos de horarios
    for _, row in df.iterrows():
        dia = row['dia']
        inicio = row['inicio']
        fin = row['fin']
        curso = f"{row['nombre_curso']} ({row['codigo_seccion']})\nProf: {row['nombre_profesor']}"

        if inicio != 'Formato inválido' and fin != 'Formato inválido':
            # Encontrar las filas correspondientes en la hoja de Excel
            for i, slot in enumerate(time_slots, start=2):  # start=2 para considerar encabezados
                next_slot = time_slots[i] if i < len(time_slots) else "19:00"
                if slot <= inicio < next_slot or slot <= fin < next_slot:
                    col = days.index(dia) + 2
                    ws.cell(row=i+1, column=col, value=curso)  # +1 para considerar encabezados
                    ws.cell(row=i+1, column=col).alignment = Alignment(wrap_text=True, vertical='center')

    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtiene la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Guardar el libro de Excel
wb.save("Horarios_Combinaciones.xlsx")
conn.close()

print("El archivo Excel 'Horarios_Combinaciones.xlsx' ha sido creado con éxito.")
